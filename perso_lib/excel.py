import os
from enum import Enum
from openpyxl import Workbook,load_workbook,cell

class ExcelMode(Enum):
    READ = 1
    WRITE = 2
    READ_WRITE = 4

class ExcelOp:
    def __init__(self,file_name,mode=ExcelMode.READ):
        self.file_name = file_name
        self.col_name_list = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'
        if mode == ExcelMode.READ:
            self.wb = load_workbook(file_name,data_only=True)

    def _get_cell_name(self,row,col):
        cell_name = self.col_name_list[col - 1] + str(row)
        return cell_name

    def open_worksheet(self,sheetname):
        if sheetname in self.wb.sheetnames:
            self.current_sheet = self.wb[sheetname]
            return True if self.current_sheet else False
        return False

    def read_cell_value(self,row,col):
        cell_name = self._get_cell_name(row,col)
        item = self.current_sheet[cell_name]
        ret = self.current_sheet[cell_name].value
        return ret

    def read_cell(self,cell_name):
        return self.current_sheet[cell_name].value